from flask import Flask, request, jsonify
import json
import os
import time
from openpyxl import load_workbook
import jwt
import base64
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import padding as asymmetric_padding
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend


app = Flask(__name__) #Flask is used for all the servers, a framewokrk for handling web applications that use HTTP such as OAuth 2.0

def generate_random_key():
    # Generate a 256-bit (32 bytes) random key
    key = os.urandom(32)
    return key

def rsa_encrypt(data, public_key): #function for encrypting data using a provided public key.
    # Encrypt data using RSA public key
    encrypted_data = public_key.encrypt(
        data,
        asymmetric_padding.OAEP(
            mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
    )
    return encrypted_data

def aes_encrypt(data, key): #function for encrypting data with AES key.
    # Generate a random IV (Initialization Vector)
    iv = os.urandom(16)

    # Create an AES-CBC cipher with the provided key and IV
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()

    # Pad the data using PKCS7
    padder = padding.PKCS7(128).padder()
    padded_data = padder.update(data) + padder.finalize()

    # Encrypt the padded data
    ciphertext = encryptor.update(padded_data) + encryptor.finalize()

    # Return the IV and the ciphertext
    return iv, ciphertext

resourceserver_private_key = rsa.generate_private_key(
    public_exponent=65537,
    key_size=2048,
)
resourceserver_public_key = resourceserver_private_key.public_key()
with open('resourceserver_private_key.pem', 'wb') as f:
    # Serialize the private key
    resourceserver_private_key_bytes = resourceserver_private_key.private_bytes(
    encoding=serialization.Encoding.PEM,
    format=serialization.PrivateFormat.PKCS8,
    encryption_algorithm=serialization.NoEncryption()
    )
    f.write(resourceserver_private_key_bytes)
with open('resourceserver_public_key.pem', 'wb') as f:
    # Serialize the private key
    resourceserver_public_key_bytes = resourceserver_public_key.public_bytes(
    encoding=serialization.Encoding.PEM,
    format=serialization.PublicFormat.SubjectPublicKeyInfo
    )
    f.write(resourceserver_public_key_bytes)

print(resourceserver_private_key_bytes)
print(resourceserver_public_key_bytes)

resourceserver_public_key = base64.b64encode(resourceserver_public_key_bytes).decode('utf-8')

@app.route('/public-key', methods=['GET'])
def get_public_key():
    return jsonify({'public_key': resourceserver_public_key})

# Function to load patient data from Excel file
def load_staff_data(filename):
    wb = load_workbook(filename)
    ws = wb.active
    staff = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        staffmember = {
            'title': row[0],
            'first_name': row[1],
            'last_name': row[2],
            'email': row[3],
            'job': row[4],
            'location': row[5],
            'clientID': row[6],
        }
        staff.append(staffmember)
    return staff

## Secret key used to sign and verify JWT access tokens
SECRET_KEY = "f4b3b8e99f43d4565ed7724a54585bde06d604f10daea12555f3b9a0a9f20e8"  # Replace with a secure secret key

# Endpoint to serve chat log files
@app.route('/chatlogs/<string:requester_name_path>/<string:client_id>', methods=['GET'])
def get_chat_log(client_id, requester_name_path):
    logfile = (f'C:/Users/Benes/PycharmProjects/glyph_messenger/chatlogs/{requester_name_path}/{client_id}.txt')
    receiverlogfile = (f'C:/Users/Benes/PycharmProjects/glyph_messenger/chatlogs/{client_id}/{requester_name_path}.txt')
    data = request.json  # Assuming the request contains JSON data

    # Create directories if they don't exist
    os.makedirs(os.path.dirname(logfile), exist_ok=True)
    os.makedirs(os.path.dirname(receiverlogfile), exist_ok=True)

    # Check if the request includes a valid access token
    access_token = request.headers.get('Authorization')
    if not access_token or not access_token.startswith('Bearer '):
        return jsonify({'error': 'invalid_token', 'description': 'Access token is missing or invalid'}), 401

    user_public_key_base64 = data.get('user_public_key')
    user_public_key = base64.b64decode(user_public_key_base64)
    print(user_public_key)
    user_public_key = serialization.load_pem_public_key(user_public_key, backend=default_backend())
    access_token = access_token.split(' ')[1]
    try:
        #decrypt acesss token using resource server private key:
        access_token = base64.b64decode(access_token)
        access_token = resourceserver_private_key.decrypt(
            access_token,
            asymmetric_padding.OAEP(
            mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
        ).decode('utf-8')
        # Decode and verify the access token
        payload = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])

        # Check expiration time
        if 'exp' in payload and payload['exp'] < int(time.time()):
            return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401

        # Access token is valid and hasn't expired
        if not os.path.exists(logfile):
            # If the file doesn't exist, create it with an empty messages list
            with open(logfile, 'w') as file:
                file.write('{"messages": []}')
        if not os.path.exists(receiverlogfile):
            with open(receiverlogfile, 'w') as file:
                file.write('{"messages": []}')
        try:
            # Read the contents of the text file
            with open(logfile, 'rb') as file:
                file_contents = file.read()
            # Generate a random AES key
            aes_key = generate_random_key()
            # Encrypt file_contents using AES-CBC
            iv, encrypted_file_contents = aes_encrypt(file_contents, aes_key)
            # Encrypt the AES key using RSA public key
            encrypted_aes_key = rsa_encrypt(aes_key, user_public_key)
            # Encode the encrypted data (base64 for easy transmission in JSON)
            encoded_iv = base64.b64encode(iv).decode('utf-8')
            encoded_encrypted_file_contents = base64.b64encode(encrypted_file_contents).decode('utf-8')
            encoded_encrypted_aes_key = base64.b64encode(encrypted_aes_key).decode('utf-8')
            # Construct JSON object
            json_data = {
                'encrypted_file_contents': encoded_encrypted_file_contents,
                'iv': encoded_iv,
                'encrypted_aes_key': encoded_encrypted_aes_key
            }
            # Print or return JSON data
            return json.dumps(json_data)
        except Exception as e:
            # Return error message if file cannot be read or parsed
            return str(e), 404
    except jwt.ExpiredSignatureError:
        # Token has expired
        return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401
    except jwt.InvalidTokenError:
        # Token is invalid
        return jsonify({'error': 'invalid_token', 'description': 'Access token is invalid'}), 401

# Endpoint to receive and append messages to chat log files
@app.route('/chatlogs/<requester_name_path>/<client_id>', methods=['POST'])
def append_to_chat_log(requester_name_path, client_id):
    try:
        # Check if the request includes a valid access token
        access_token = request.headers.get('Authorization')
        if not access_token or not access_token.startswith('Bearer '):
            return jsonify({'error': 'invalid_token', 'description': 'Access token is missing or invalid'}), 401
        #decrypt acesss token using resource server private key:
        access_token = access_token.split(' ')[1]
        access_token = base64.b64decode(access_token)
        access_token = resourceserver_private_key.decrypt(
            access_token,
            asymmetric_padding.OAEP(
            mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
        ).decode('utf-8')
        # Decode and verify the access token
        payload = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])
        # Check expiration time
        if 'exp' in payload and payload['exp'] < int(time.time()):
            return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401

        # Get the text message data from the request
        encrypted_message_bytes = request.get_data()
        encrypted_message_bytes = base64.b64decode(encrypted_message_bytes)
        # Decrypt the encrypted message using the private key
        decrypted_message_bytes = resourceserver_private_key.decrypt(
            encrypted_message_bytes,
            asymmetric_padding.OAEP(
                mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
        decrypted_message_str = decrypted_message_bytes.decode('utf-8')
        message_data  = json.loads(decrypted_message_str)

        # Construct the path to the requester's chat log file
        requester_chat_log_file_path = os.path.join('chatlogs', requester_name_path, f'{client_id}.txt')

        # Construct the path to the client's chat log file
        client_chat_log_file_path = os.path.join('chatlogs', client_id, f'{requester_name_path}.txt')

        # Check if the requester's chat log file exists
        if os.path.exists(requester_chat_log_file_path):
            # Open the requester's chat log file in append mode and write the new message
            with open(requester_chat_log_file_path, 'r') as file:
                data = json.load(file)
            if isinstance(message_data, str):
                message_data = json.loads(message_data)
            # Append new message to the "messages" list
            data['messages'].append(message_data)
            # Write updated JSON data back to file
            with open(requester_chat_log_file_path, 'w') as file:
                json.dump(data, file, indent=2)

        # Check if the client's chat log file exists
        if os.path.exists(client_chat_log_file_path):
            # Open the client's chat log file in append mode and write the new message
            with open(client_chat_log_file_path, 'r') as file:
                data = json.load(file)
            if isinstance(message_data, str):
                message_data = json.loads(message_data)
            # Append new message to the "messages" list
            data['messages'].append(message_data)
            # Write updated JSON data back to file
            with open(client_chat_log_file_path, 'w') as file:
                json.dump(data, file, indent=2)

        return jsonify({'message': 'Message appended to chat logs successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/newchats/<requester_name_path>/', methods=['GET'])
def get_file_sizes(requester_name_path): #checking file siez to see if any chatlog file has increased in size, denoting a new message received.
    # Check if the request includes a valid access token
    access_token = request.headers.get('Authorization')
    if not access_token or not access_token.startswith('Bearer '):
        return jsonify({'error': 'invalid_token', 'description': 'Access token is missing or invalid'}), 401
    access_token = access_token.split(' ')[1]
    try:
        #decrypt acesss token using resource server private key:
        access_token = base64.b64decode(access_token)
        access_token = resourceserver_private_key.decrypt(
            access_token,
            asymmetric_padding.OAEP(
            mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
        ).decode('utf-8')
        # Decode and verify the access token
        payload = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])
        # Check expiration time
        if 'exp' in payload and payload['exp'] < int(time.time()):
            return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401

        # Directory path for chat logs
        chatlogs_directory = f'C:/Users/Benes/PycharmProjects/glyph_messenger/chatlogs/{requester_name_path}/'

        # Read file sizes from the config file
        config_file_path = os.path.join(chatlogs_directory, 'config.txt')

        # Create the config file if it doesn't exist
        if not os.path.exists(config_file_path):
            with open(config_file_path, 'w') as config_file:
                json.dump({}, config_file)

        # Read old file sizes from the config file
        old_file_sizes = {}
        if os.path.exists(config_file_path):
            with open(config_file_path, 'r') as config_file:
                old_file_sizes = json.load(config_file)

        # Get the list of files in the specified directory
        files = os.listdir(chatlogs_directory)

        # Dictionary to store updated file sizes
        updated_file_sizes = {}

        # Signal any changes to file sizes
        changes = {}

        # Iterate over each file in the directory
        for file_name in files:
            if file_name == 'config.txt':
                continue
            file_path = os.path.join(chatlogs_directory, file_name)
            # Get the size of the file
            size = os.path.getsize(file_path)
            # Store the size in the dictionary
            updated_file_sizes[file_name] = size

            # Check if file is new or its size has increased
            if file_name not in old_file_sizes or old_file_sizes.get(file_name) != size:
                changes[file_name] = {
                    'size': size,
                    'increased': size > old_file_sizes.get(file_name, 0)
                }

        # Update the config file with the latest file sizes
        with open(config_file_path, 'w') as config_file:
            json.dump(updated_file_sizes, config_file, indent=2)

        # Return full list of file sizes and changes as JSON response
        return jsonify({'file_sizes': updated_file_sizes, 'changes': changes}), 200

    except Exception as e:
        # Return error message if an exception occurs
        return jsonify({'error': str(e)}), 500

@app.route('/chathistory/<requester_name_path>/', methods=['GET']) #get chat history for the requseting client, to fill their contact list box.
def get_chat_history(requester_name_path):
    try:
        # Directory path for chat logs
        chatlogs_directory = f'C:/Users/Benes/PycharmProjects/glyph_messenger/chatlogs/{requester_name_path}/'
        # Get the list of files in the specified directory
        files = os.listdir(chatlogs_directory)
        filenames_without_extension = [os.path.splitext(file)[0] for file in files]
        # Return the list of filenames as JSON response
        return jsonify(filenames_without_extension), 200
    except Exception as e:
        # Return error message if an exception occurs
        return jsonify({'error': str(e)}), 500

@app.route('/staffdata', methods=['GET']) #providing staff data based on valid access tokens.
def get_staffdata():
    # Check if the request includes a valid access token
    access_token = request.headers.get('Authorization')
    if not access_token or not access_token.startswith('Bearer '):
        return jsonify({'error': 'invalid_token', 'description': 'Access token is missing or invalid'}), 401
    access_token = access_token.split(' ')[1]
    try:
        #decrypt acesss token using resource server private key:
        access_token = base64.b64decode(access_token)
        access_token = resourceserver_private_key.decrypt(
            access_token,
            asymmetric_padding.OAEP(
            mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
        ).decode('utf-8')
        # Decode and verify the access token
        payload = jwt.decode(access_token, SECRET_KEY, algorithms=['HS256'])
        # Check expiration time
        if 'exp' in payload and payload['exp'] < int(time.time()):
            return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401

        # Access token is valid and hasn't expired
        staff = load_staff_data('staff.xlsx')
        return jsonify({'staff': staff})

    except jwt.ExpiredSignatureError:
        # Token has expired
        return jsonify({'error': 'invalid_token', 'description': 'Access token has expired'}), 401
    except jwt.InvalidTokenError:
        # Token is invalid
        return jsonify({'error': 'invalid_token', 'description': 'Access token is invalid'}), 401


def validate_access_token(access_token): #function for validating access tokens
    try:
        # Decode the access token using the secret key
        decoded_token = jwt.decode(access_token, SECRET_KEY, algorithms=["HS256"])
        return True
    except jwt.ExpiredSignatureError:
        print("Token has expired")
    except jwt.InvalidTokenError:
        print("Invalid token")
    return False

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
